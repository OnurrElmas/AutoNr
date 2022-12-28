import tkinter as tk    
from tkinter import ttk
from tkinter import *
import pandas as pd
import openpyxl


def connect_with_excel():
        
    path = "C:\\Users\\ONURELMA\\Desktop\PDFs\\pandas.xlsm"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    cellstrings = []
    cellnumbers = []
    for i in range (1,100):
        cell_obj = sheet_obj.cell(row = 1, column = i).value
        cellstring = str(cell_obj)
        if cellstring.startswith('SW_'):
            cellstrings.append(cellstring)
    return cellstrings


class ComboboxSelectionWindow():
    def __init__(self, master):
        self.master=master
        self.entry_contents=None
        self.labelTop = tk.Label(master,text = "Select one of the SW Version")
        self.labelTop.place(x = 30, y = 10, width=140, height=10)
        self.comboBox_example = ttk.Combobox(master,values=cellstrings)
        self.comboBox_example.current(0)
        self.comboBox_example.place(x = 20, y = 30, width=140, height=25)

        self.okButton = tk.Button(master, text='OK',command = self.callback)
        self.okButton.place(x = 20, y = 60, width=140, height=25)

    def callback(self):
        """ get the contents of the Entry and exit
        """
        self.comboBox_example_contents=self.comboBox_example.get()
        self.master.destroy()
        
def ComboboxSelection():

    app = tk.Tk()
    app.geometry('500x250')
    Selection=ComboboxSelectionWindow(app)
    app.mainloop()

    #print("Selected interface: ", Selection.comboBox_example_contents)
    return Selection.comboBox_example_contents


class CheckBoxSelectionWindow():
    def __init__(self, master):
        self.master=master
        self.entry_contents=None
        self.chvars = []

        self.chvar_1 = tk.IntVar()
        self.chvar_1.set(0)
        self.checkbox_example = ttk.Checkbutton(master,text="Integration",variable=self.chvar_1).grid(row=0,column=0)
        

        self.chvar_2 = tk.IntVar()
        self.chvar_2.set(0)
        self.checkbox_example = ttk.Checkbutton(master,text="Fusi Prio 1",variable=self.chvar_2).grid(row=10,column=0)
        

        self.chvar_3 = tk.IntVar()
        self.chvar_3.set(0)
        self.checkbox_example = ttk.Checkbutton(master,text="Fusi Prio 2",variable=self.chvar_3).grid(row=20,column=0)

        self.chvar_4 = tk.IntVar()
        self.chvar_4.set(0)
        self.checkbox_example = ttk.Checkbutton(master,text="Fusi Prio 3",variable=self.chvar_4).grid(row=30,column=0)

        self.chvar_5 = tk.IntVar()
        self.chvar_5.set(0)
        self.checkbox_example = ttk.Checkbutton(master,text="QM Prio 1",variable=self.chvar_5).grid(row=40,column=0)

        self.chvar_6 = tk.IntVar()
        self.chvar_6.set(0)
        self.checkbox_example = ttk.Checkbutton(master,text="QM Prio 2",variable=self.chvar_6).grid(row=50,column=0)
        
        self.chvar_7 = tk.IntVar()
        self.chvar_7.set(0)
        self.checkbox_example = ttk.Checkbutton(master,text="QM Prio 3",variable=self.chvar_7).grid(row=60,column=0)
        
        self.okButton = tk.Button(master, text='OK',command = self.callback)
        self.okButton.place(x = 100, y = 100, width=140, height=25)

    def callback(self):
        """ get the contents of the Entry and exit
        """
        #print(self.chvar.get())
        self.master.destroy()

def CheckBoxSelection():

    app = tk.Tk()
    app.geometry('500x250')
    Selection=CheckBoxSelectionWindow(app)
    app.mainloop()
    chvars = []
    chvars.append(Selection.chvar_1.get())
    chvars.append(Selection.chvar_2.get())
    chvars.append(Selection.chvar_3.get())
    chvars.append(Selection.chvar_4.get())
    chvars.append(Selection.chvar_5.get())
    chvars.append(Selection.chvar_6.get())
    chvars.append(Selection.chvar_7.get())
    #print("Selected interface: ", Selection.comboBox_example_contents)

    return chvars

def find_column(comboselection):
    path = "C:\\Users\\ONURELMA\\Desktop\\PDFs\\pandas.xlsm"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    cellstrings = []
    for i in range (1,100):
        cell_obj = sheet_obj.cell(row = 1, column = i).value
        cellstring = str(cell_obj)
        if cellstring == comboselection:
            h=i
    return h
    
def main_function(columnnumber,chvars):
    path = "C:\\Users\\ONURELMA\\Desktop\\PDFs\\pandas.xlsm"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    cellstrings = []
    for i in range (3,700):
        cell_obj = sheet_obj.cell(row = i, column = columnnumber).value
        cellstring = str(cell_obj)
        if (cell_obj is None or cellstring == 'R' or cellstring == 'O') and sheet_obj.cell(row = i, column = 13).value == '1M' and sheet_obj.cell(row = i, column = 18).value is None:
            cellstrings.append(sheet_obj.cell(row = i, column = 10).value)
    return cellstrings
    

def write_in_txt(tc_numbers):
    with open("C:\\Users\\ONURELMA\\Desktop\\PDFs\\NightRun.txt","w") as file1:
        for tc_number in tc_numbers:
            file1.write("%s\n" %tc_number)
    file1.close
    

#eger chvars in 0 inci elemani 1 ise and String = colum ise


cellstrings = connect_with_excel()

comboselection = ComboboxSelection()
checkboxselection = CheckBoxSelection()
print(comboselection)
print(checkboxselection)

columnnumber = find_column(comboselection)
print(columnnumber)
tc_numbers =main_function(columnnumber,checkboxselection)
write_in_txt(tc_numbers)
